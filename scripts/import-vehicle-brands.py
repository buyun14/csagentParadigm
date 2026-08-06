#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
车型库导入脚本：将 docs/车型库0602(1).xlsx 解析为 src/lib/agent/vehicle-brands.generated.ts

用法（Windows PowerShell）：
    python scripts/import-vehicle-brands.py

依赖：Python 3 + openpyxl（`pip install openpyxl`）

处理规则：
1. 跳过表头行（品牌/车系）
2. 品牌名归一化（BRAND_MERGES：'理想汽车' → '理想'，与 knowledge-base.ts 的 brandMerges 保持一致）
3. 车系名去除品牌前缀（'蔚来EC6' → 'EC6'；去前缀后为空则保留原名，如 MINI 的 'MINI'）
4. 输出按品牌排序的 TS 文件，文件头记录来源与生成时间

注意：与 src/lib/agent/knowledge-base.ts 中 manualSeries（type/power/priceRange 附加字段）
与 brandMerges 保持同步；本脚本只负责品牌/车系名数据。
"""

import datetime
import os
import openpyxl

XLSX_PATH = os.path.join('docs', '车型库0602(1).xlsx')
OUT_PATH = os.path.join('src', 'lib', 'agent', 'vehicle-brands.generated.ts')

# 与 knowledge-base.ts 的 brandMerges 保持一致
BRAND_MERGES = {'理想汽车': '理想'}


def normalize_series(brand: str, series: str) -> str:
    """按合并后品牌名去除车系前缀；去前缀为空则保留原名"""
    eff = BRAND_MERGES.get(brand, brand)
    if len(eff) >= 2 and series.startswith(eff):
        candidate = series[len(eff):].strip()
        if candidate:
            return candidate
    return series


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    brands: dict[str, set[str]] = {}
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        b = str(row[0]).strip()
        s = str(row[1]).strip()
        if not header_seen and b == '品牌' and s == '车系':
            header_seen = True
            continue
        if not b or not s:
            continue
        s = normalize_series(b, s)
        brands.setdefault(b, set()).add(s)

    assert '品牌' not in brands, '表头行未跳过，请检查 xlsx 结构'

    total_series = sum(len(v) for v in brands.values())
    lines = [
        '// 由 docs/车型库0602(1).xlsx 自动生成（scripts/import-vehicle-brands.py，品牌前缀已归一化去除）',
        '// 生成时间: ' + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        '// 数据规模: 品牌 %d / 车系 %d。仅含品牌与车系名，无车身类型/动力/价格字段（见 knowledge-base.ts 合并逻辑）'
        % (len(brands), total_series),
        'export const vehicleBrandSeries: Record<string, string[]> = {',
    ]
    for b in sorted(brands):
        lines.append('  %r: [%s],' % (b, ', '.join(repr(s) for s in sorted(brands[b]))))
    lines.append('};')

    with open(OUT_PATH, 'w', encoding='utf-8', newline='\n') as f:
        f.write('\n'.join(lines) + '\n')
    print('written: %s (%d brands / %d series)' % (OUT_PATH, len(brands), total_series))


if __name__ == '__main__':
    main()
